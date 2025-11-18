"""
Reports Service
Business logic for generating reports in various formats
"""
import io
import csv
import json
from typing import BinaryIO, Dict, Any, List
from uuid import UUID, uuid4
from datetime import datetime, timedelta
from sqlalchemy.orm import Session
from loguru import logger

from modules.reports.schemas import (
    ReportRequest,
    ReportResponse,
    ReportFormat,
    ReportType,
    TimePeriod
)
from modules.analytics.service import AnalyticsService
from modules.analytics.schemas import AnalyticsQueryParams


class ReportService:
    """Service for generating reports"""

    def __init__(self, db: Session):
        self.db = db
        self.analytics_service = AnalyticsService(db)

    # ==================== Report Generation ====================

    async def generate_report(self, request: ReportRequest) -> ReportResponse:
        """
        Generate a report in the specified format

        Process:
          1. Fetch analytics data
          2. Generate report in requested format
          3. Save to storage
          4. Return download URL

        Args:
            request: Report generation request

        Returns:
            ReportResponse with download URL
        """
        # Fetch analytics data
        data = await self._fetch_report_data(request)

        # Generate report in requested format
        if request.format == ReportFormat.CSV:
            file_content = await self._generate_csv(data, request)
            extension = "csv"
            mime_type = "text/csv"

        elif request.format == ReportFormat.EXCEL:
            file_content = await self._generate_excel(data, request)
            extension = "xlsx"
            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        elif request.format == ReportFormat.PDF:
            file_content = await self._generate_pdf(data, request)
            extension = "pdf"
            mime_type = "application/pdf"

        else:
            raise ValueError(f"Unsupported format: {request.format}")

        # Generate filename
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"{request.report_type.value}_{timestamp}.{extension}"

        # In production, save to S3/cloud storage
        # For now, return a mock response
        report_id = uuid4()
        download_url = f"/api/v1/reports/download/{report_id}"

        return ReportResponse(
            id=report_id,
            report_type=request.report_type.value,
            format=request.format.value,
            filename=filename,
            file_size_bytes=len(file_content),
            download_url=download_url,
            generated_at=datetime.utcnow(),
            expires_at=datetime.utcnow() + timedelta(days=7)
        )

    async def _fetch_report_data(self, request: ReportRequest) -> Dict[str, Any]:
        """Fetch analytics data for the report"""

        # Build analytics query params
        params = AnalyticsQueryParams(
            tenant_id=request.tenant_id,
            time_period=request.time_period,
            start_date=request.start_date,
            end_date=request.end_date,
            include_breakdown=request.include_breakdown,
            practitioner_id=request.practitioner_id,
            department=request.department,
            location_id=request.location_id
        )

        # Fetch appropriate data based on report type
        if request.report_type == ReportType.APPOINTMENTS:
            result = self.analytics_service.get_appointment_analytics(params)
            return {
                "type": "appointments",
                "summary": result.summary,
                "breakdown": result.breakdown,
                "time_period": result.time_period,
                "start_date": result.start_date,
                "end_date": result.end_date
            }

        elif request.report_type == ReportType.JOURNEYS:
            result = self.analytics_service.get_journey_analytics(params)
            return {
                "type": "journeys",
                "summary": result.summary,
                "breakdown": result.breakdown,
                "time_period": result.time_period,
                "start_date": result.start_date,
                "end_date": result.end_date
            }

        elif request.report_type == ReportType.COMMUNICATION:
            result = self.analytics_service.get_communication_analytics(params)
            return {
                "type": "communication",
                "summary": result.summary,
                "breakdown": result.breakdown,
                "time_period": result.time_period,
                "start_date": result.start_date,
                "end_date": result.end_date
            }

        elif request.report_type == ReportType.VOICE_CALLS:
            result = self.analytics_service.get_voice_call_analytics(params)
            return {
                "type": "voice_calls",
                "summary": result.summary,
                "breakdown": result.breakdown,
                "time_period": result.time_period,
                "start_date": result.start_date,
                "end_date": result.end_date
            }

        elif request.report_type == ReportType.DASHBOARD:
            result = self.analytics_service.get_dashboard_overview(
                request.tenant_id,
                request.time_period,
                request.start_date,
                request.end_date
            )
            return {
                "type": "dashboard",
                "appointments": result.appointments,
                "journeys": result.journeys,
                "communication": result.communication,
                "voice_calls": result.voice_calls,
                "time_period": result.time_period,
                "start_date": result.start_date,
                "end_date": result.end_date
            }

        else:
            raise ValueError(f"Unsupported report type: {request.report_type}")

    # ==================== CSV Generation ====================

    async def _generate_csv(self, data: Dict[str, Any], request: ReportRequest) -> bytes:
        """Generate CSV report"""
        output = io.StringIO()

        if request.report_type == ReportType.APPOINTMENTS:
            writer = csv.writer(output)

            # Header
            writer.writerow(["Appointment Analytics Report"])
            writer.writerow([f"Period: {data['time_period']}"])
            writer.writerow([f"From: {data['start_date']} To: {data['end_date']}"])
            writer.writerow([])

            # Summary metrics
            writer.writerow(["Summary Metrics"])
            writer.writerow(["Metric", "Value"])
            summary = data['summary']
            writer.writerow(["Total Appointments", summary.total_appointments])
            writer.writerow(["Scheduled", summary.scheduled])
            writer.writerow(["Confirmed", summary.confirmed])
            writer.writerow(["Completed", summary.completed])
            writer.writerow(["Canceled", summary.canceled])
            writer.writerow(["No Show", summary.no_show])
            writer.writerow(["Completion Rate", f"{summary.completion_rate:.2f}%"])
            writer.writerow(["No Show Rate", f"{summary.no_show_rate:.2f}%"])
            writer.writerow(["Cancellation Rate", f"{summary.cancellation_rate:.2f}%"])
            writer.writerow([])

            # Trend data
            if request.include_trends and summary.trend:
                writer.writerow(["Daily Trend"])
                writer.writerow(["Date", "Appointments"])
                for point in summary.trend:
                    writer.writerow([point.date, point.value])
                writer.writerow([])

            # Breakdown
            if request.include_breakdown and data['breakdown']:
                breakdown = data['breakdown']

                if breakdown.by_channel:
                    writer.writerow(["By Channel"])
                    writer.writerow(["Channel", "Count"])
                    for channel, count in breakdown.by_channel.items():
                        writer.writerow([channel, count])
                    writer.writerow([])

                if breakdown.by_practitioner:
                    writer.writerow(["By Practitioner"])
                    writer.writerow(["Practitioner", "Count"])
                    for practitioner, count in breakdown.by_practitioner.items():
                        writer.writerow([practitioner, count])
                    writer.writerow([])

                if breakdown.by_department:
                    writer.writerow(["By Department"])
                    writer.writerow(["Department", "Count"])
                    for department, count in breakdown.by_department.items():
                        writer.writerow([department, count])
                    writer.writerow([])

        elif request.report_type == ReportType.DASHBOARD:
            writer = csv.writer(output)

            writer.writerow(["PRM Dashboard Overview"])
            writer.writerow([f"Period: {data['time_period']}"])
            writer.writerow([f"From: {data['start_date']} To: {data['end_date']}"])
            writer.writerow([])

            # Appointments
            writer.writerow(["Appointments"])
            writer.writerow(["Metric", "Value"])
            apt = data['appointments']
            writer.writerow(["Total", apt.total_appointments])
            writer.writerow(["Completion Rate", f"{apt.completion_rate:.2f}%"])
            writer.writerow([])

            # Journeys
            writer.writerow(["Journeys"])
            writer.writerow(["Metric", "Value"])
            jrn = data['journeys']
            writer.writerow(["Total Active", jrn.total_active])
            writer.writerow(["Completed", jrn.completed])
            writer.writerow(["Completion Rate", f"{jrn.completion_rate:.2f}%"])
            writer.writerow([])

            # Communication
            writer.writerow(["Communication"])
            writer.writerow(["Metric", "Value"])
            comm = data['communication']
            writer.writerow(["Total Messages", comm.total_messages])
            writer.writerow(["Total Conversations", comm.total_conversations])
            writer.writerow(["Avg Response Time (min)", f"{comm.avg_response_time_minutes:.1f}"])
            writer.writerow([])

            # Voice Calls
            writer.writerow(["Voice Calls"])
            writer.writerow(["Metric", "Value"])
            voice = data['voice_calls']
            writer.writerow(["Total Calls", voice.total_calls])
            writer.writerow(["Booking Success Rate", f"{voice.booking_success_rate:.2f}%"])
            writer.writerow(["Avg Duration (min)", f"{voice.avg_duration_minutes:.1f}"])

        else:
            # Generic CSV for other report types
            writer = csv.writer(output)
            writer.writerow(["Report Type", request.report_type.value])
            writer.writerow(["Generated At", datetime.utcnow().isoformat()])
            writer.writerow([])
            writer.writerow(["Data"])
            writer.writerow([json.dumps(data, indent=2, default=str)])

        content = output.getvalue()
        return content.encode('utf-8')

    # ==================== Excel Generation ====================

    async def _generate_excel(self, data: Dict[str, Any], request: ReportRequest) -> bytes:
        """Generate Excel report"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Fallback to CSV if openpyxl not installed
            logger.warning("openpyxl not installed, falling back to CSV")
            return await self._generate_csv(data, request)

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        # Styling
        title_font = Font(size=14, bold=True)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        row = 1

        # Title
        title = request.title or f"{request.report_type.value.title()} Report"
        ws.cell(row, 1, title).font = title_font
        row += 1
        ws.cell(row, 1, f"Period: {data.get('time_period', 'N/A')}")
        row += 1
        ws.cell(row, 1, f"From: {data.get('start_date', 'N/A')} To: {data.get('end_date', 'N/A')}")
        row += 2

        # Summary section
        if 'summary' in data and data['summary']:
            ws.cell(row, 1, "Summary Metrics").font = header_font
            row += 1

            summary = data['summary']
            for attr in dir(summary):
                if not attr.startswith('_') and not attr == 'trend':
                    value = getattr(summary, attr, None)
                    if value is not None:
                        ws.cell(row, 1, attr.replace('_', ' ').title())
                        ws.cell(row, 2, str(value))
                        row += 1

            row += 1

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # ==================== PDF Generation ====================

    async def _generate_pdf(self, data: Dict[str, Any], request: ReportRequest) -> bytes:
        """Generate PDF report"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
        except ImportError:
            # Fallback to CSV if reportlab not installed
            logger.warning("reportlab not installed, falling back to CSV")
            return await self._generate_csv(data, request)

        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title = request.title or f"{request.report_type.value.title()} Report"
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a56db'),
            alignment=TA_CENTER
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.3 * inch))

        # Metadata
        meta_text = f"""
        <b>Period:</b> {data.get('time_period', 'N/A')}<br/>
        <b>From:</b> {data.get('start_date', 'N/A')}<br/>
        <b>To:</b> {data.get('end_date', 'N/A')}<br/>
        <b>Generated:</b> {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}
        """
        elements.append(Paragraph(meta_text, styles['Normal']))
        elements.append(Spacer(1, 0.3 * inch))

        # Summary table
        if 'summary' in data and data['summary']:
            elements.append(Paragraph("<b>Summary Metrics</b>", styles['Heading2']))
            elements.append(Spacer(1, 0.1 * inch))

            summary = data['summary']
            table_data = [["Metric", "Value"]]

            for attr in dir(summary):
                if not attr.startswith('_') and not attr == 'trend':
                    value = getattr(summary, attr, None)
                    if value is not None:
                        metric_name = attr.replace('_', ' ').title()
                        table_data.append([metric_name, str(value)])

            table = Table(table_data, colWidths=[3 * inch, 2 * inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')])
            ]))
            elements.append(table)

        # Build PDF
        doc.build(elements)
        return buffer.getvalue()


def get_report_service(db: Session) -> ReportService:
    """Dependency injection for ReportService"""
    return ReportService(db)
